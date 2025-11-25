import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelHandler:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.dataframe = None

    def read_excel(self):
        """Чтение Excel файла"""
        try:
            self.dataframe = pd.read_excel(self.file_path)
            return self.dataframe
        except Exception as e:
            print(f"Ошибка чтения файла: {e}")
            return None

    def write_excel(self, df, sheet_name='Sheet1'):
        """Запись DataFrame в Excel"""
        try:
            with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            print(f"Ошибка записи в файл: {e}")

    def create_empty_excel(self):
        """Создание пустого Excel файла"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист1"
        wb.save(self.file_path)

    def append_data(self, df, sheet_name='Sheet1'):
        """Добавление данных в существующий файл"""
        try:
            writer = pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a')
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=writer.sheets[sheet_name].max_row)
            writer.save()
        except Exception as e:
            print(f"Ошибка добавления данных: {e}")

    def get_sheet_names(self):
        """Получение списка листов в файле"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            return self.workbook.sheetnames
        except Exception as e:
            print(f"Ошибка получения названий листов: {e}")
            return []

    def delete_sheet(self, sheet_name):
        """Удаление листа"""
        try:
            wb = openpyxl.load_workbook(self.file_path)
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
                wb.save(self.file_path)
        except Exception as e:
            print(f"Ошибка удаления листа: {e}")


# Пример использования
if __name__ == "__main__":
    excel_handler = ExcelHandler("data.xlsx")

    # Создание пустого файла
    excel_handler.create_empty_excel()

    # Создание тестовых данных
    data = {
        'Имя': ['Иван', 'Петр', 'Анна'],
        'Возраст': [25, 30, 22],
        'Зарплата': [50000, 60000, 55000]
    }

    df = pd.DataFrame(data)

    # Запись данных
    excel_handler.write_excel(df)

    # Чтение данных
    read_df = excel_handler.read_excel()
    print(read_df)
